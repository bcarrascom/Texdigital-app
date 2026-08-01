"""
excel/generador_documentos.py
Exportación de cotizaciones a Excel usando openpyxl.
"""

import shutil
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    raise ImportError("Instala openpyxl: pip install openpyxl")

# ── Rutas base ─────────────────────────────────────────────────────────────────
from core.rutas import RECURSOS
PLANTILLA = RECURSOS / "Cotizador Backlight.xlsx"

# Carpeta de descargas según plataforma
def _carpeta_descargas() -> Path:
    import sys, os
    if sys.platform == "win32":
        return Path.home() / "Downloads"
    elif sys.platform == "darwin":
        return Path.home() / "Downloads"
    else:
        # Linux: intentar XDG_DOWNLOAD_DIR, si no ~/Downloads
        xdg = os.environ.get("XDG_DOWNLOAD_DIR")
        return Path(xdg) if xdg else Path.home() / "Downloads"


# ── Mapeo de columnas por campo de producto ────────────────────────────────────
# Fila base de productos: 10 (B10..B110 = filas 10..110, hasta 101 productos)
FILA_INI_PRODUCTOS = 10
FILA_FIN_PRODUCTOS = 110

# Columna letra → campo del dict de producto
COL_PRODUCTO  = "B"   # nombre/código (usaremos el tema como identificador)
COL_TEXTIL    = "C"   # tela
COL_ANCHO     = "E"   # ancho (m)
COL_ALTO      = "F"   # alto (m)
COL_CANTIDAD  = "G"   # cantidad
COL_TEMA      = "L"   # tema


def exportar_cotizacion_backlight(
    datos_cliente: dict,
    datos_productos: list[dict],
    destino_dir: Path | None = None,
) -> str:
    """
    Crea una copia de la plantilla, rellena los datos y la guarda
    en destino_dir (si se pasa) o en la carpeta de Descargas.
    Devuelve la ruta del archivo generado como string.
    """

    if not PLANTILLA.exists():
        raise FileNotFoundError(
            f"No se encontró la plantilla en:\n{PLANTILLA}\n"
            "Asegúrate de que 'Cotizador Backlight.xlsx' está en la carpeta del programa."
        )

    # ── Crear copia ────────────────────────────────────────────────────────────
    cot = datos_cliente.get("cotizacion", "0000")
    nombre_archivo = f"Cotización {cot}.xlsx"
    carpeta = destino_dir if destino_dir is not None else _carpeta_descargas()
    destino = carpeta / nombre_archivo
    shutil.copy2(PLANTILLA, destino)

    wb = openpyxl.load_workbook(destino)

    # Localizar la hoja "Nota de venta" (búsqueda insensible a mayúsculas)
    nombre_hoja = next(
        (n for n in wb.sheetnames if n.lower() == "nota de venta"),
        None,
    )
    if nombre_hoja is None:
        raise ValueError(
            f"No se encontró la hoja 'Nota de venta' en la plantilla.\n"
            f"Hojas disponibles: {', '.join(wb.sheetnames)}"
        )
    ws = wb[nombre_hoja]

    # ── Datos del cliente ──────────────────────────────────────────────────────
    _fecha_raw = datos_cliente.get("fecha", "")
    try:
        _fecha_dt = datetime.strptime(_fecha_raw, "%d/%m/%Y")
    except Exception:
        _fecha_dt = _fecha_raw
    _set(ws, "C2", _fecha_dt)
    _set(ws, "H1",  datos_cliente.get("cotizacion",   ""))
    _set(ws, "C5",  datos_cliente.get("contacto",     ""))
    _set(ws, "C6",  datos_cliente.get("empresa",      ""))
    _set(ws, "C7",  datos_cliente.get("email",        ""))
    _set(ws, "I5",  datos_cliente.get("razon_social", ""))
    _set(ws, "I6",  datos_cliente.get("rut",          ""))
    _descuento_raw = datos_cliente.get("descuento", "")
    try:
        _descuento_val = float(_descuento_raw) / 100 if _descuento_raw != "" else ""
    except (ValueError, TypeError):
        _descuento_val = _descuento_raw
    _set(ws, "I7", _descuento_val)
    _set(ws, "C19",  datos_cliente.get("descripcion",   ""))
    _set(ws, "C112", datos_cliente.get("nombre_trabajo",""))
    _set(ws, "C113", datos_cliente.get("descripcion",   ""))

    # ── Productos (filas 10..17) ───────────────────────────────────────────────
    for offset, d in enumerate(datos_productos):
        fila = FILA_INI_PRODUCTOS + offset
        if fila > FILA_FIN_PRODUCTOS:
            break

        caja = d.get("caja", "Sin caja")
        producto = "Caja + Backlight" if caja and caja != "Sin caja" else "Textil Backlight"
        _set(ws, f"{COL_PRODUCTO}{fila}", producto)
        _set(ws, f"{COL_TEXTIL}{fila}",   _limpiar_texto(d.get("tela",  "")))
        _set(ws, f"{COL_ANCHO}{fila}",    d.get("ancho",    ""))
        _set(ws, f"{COL_ALTO}{fila}",     d.get("alto",     ""))
        _set(ws, f"{COL_CANTIDAD}{fila}", d.get("cantidad", ""))
        _set(ws, f"{COL_TEMA}{fila}",     _limpiar_texto(d.get("tema", "")))

    # Filas sobrantes: ocultar y limpiar
    for fila in range(FILA_INI_PRODUCTOS + len(datos_productos),
                      FILA_FIN_PRODUCTOS + 1):
        for col in [COL_PRODUCTO, COL_TEXTIL, COL_ANCHO, COL_ALTO,
                    COL_CANTIDAD, COL_TEMA]:
            _set(ws, f"{col}{fila}", "")
    _ocultar_sobrantes(ws, FILA_INI_PRODUCTOS, FILA_FIN_PRODUCTOS, len(datos_productos))

    # ── Hoja Cotización ───────────────────────────────────────────────────────
    nombre_cotizacion = next(
        (n for n in wb.sheetnames
         if n.lower() in ("cotización", "cotizacion")),
        None,
    )
    if nombre_cotizacion:
        ws_cot = wb[nombre_cotizacion]
        _set(ws_cot, "C7", datos_cliente.get("nombre_trabajo", ""))
        _set(ws_cot, "H7", datos_cliente.get("condicion", ""))
        _ocultar_sobrantes(ws_cot, 12, 111, len(datos_productos))

    # ── Hoja OP — se elimina en cotizaciones Backlight ───────────────────────
    nombre_op = next(
        (n for n in wb.sheetnames if n.lower() == "op"),
        None,
    )
    if nombre_op:
        wb.remove(wb[nombre_op])

    # ── Hoja OP TELAS ─────────────────────────────────────────────────────────
    nombre_op_telas = next(
        (n for n in wb.sheetnames if n.lower() == "op telas"),
        None,
    )
    if nombre_op_telas:
        ws_op_telas = wb[nombre_op_telas]
        _set(ws_op_telas, "J7", datos_cliente.get("terminaciones_caja", "CAJA TERMINADA"))
        _ocultar_sobrantes(ws_op_telas, 9, 108, len(datos_productos))

    # ── Hoja NV CAJAS (BL) — perfiles C16..C23 ────────────────────────────────
    nombre_nv_cajas = next(
        (n for n in wb.sheetnames if n.lower() == "nv cajas (bl)"),
        None,
    )
    if nombre_nv_cajas:
        ws_nv = wb[nombre_nv_cajas]
        FILA_INI_NV = 16
        for offset, d in enumerate(datos_productos):
            fila = FILA_INI_NV + offset
            if fila > 115:
                break
            caja = d.get("caja", "")
            _set(ws_nv, f"C{fila}", caja if caja and caja != "Sin caja" else "")

    # ── Abrir en hoja Cotización ───────────────────────────────────────────────
    _activar_hoja_cotizacion(wb)

    wb.save(destino)
    wb.close()

    return str(destino)


# Hojas exclusivas de backlight que se eliminan en cotizaciones no-backlight
_HOJAS_SOLO_BACKLIGHT = {"nv textil (bl)", "op cajas", "op telas"}


def exportar_cotizacion_nueva(
    datos_cliente: dict,
    datos_productos: list[dict],
    destino_dir: Path | None = None,
) -> str:
    """
    Crea una cotización no-backlight desde la misma plantilla,
    eliminando las hojas exclusivas de backlight y rellenando
    producto/textil en lugar de caja/tela.
    """
    if not PLANTILLA.exists():
        raise FileNotFoundError(
            f"No se encontró la plantilla en:\n{PLANTILLA}\n"
            "Asegúrate de que 'Cotizador Backlight.xlsx' está en la carpeta del programa."
        )

    cot = datos_cliente.get("cotizacion", "0000")
    carpeta = destino_dir if destino_dir is not None else _carpeta_descargas()
    destino = carpeta / f"Cotización {cot}.xlsx"
    shutil.copy2(PLANTILLA, destino)

    wb = openpyxl.load_workbook(destino)

    # ── Eliminar hojas exclusivas de backlight ─────────────────────────────────
    for nombre in list(wb.sheetnames):
        if nombre.lower() in _HOJAS_SOLO_BACKLIGHT:
            wb.remove(wb[nombre])

    # ── Ocultar NV CAJAS (BL) ──────────────────────────────────────────────────
    nombre_nv_cajas = next(
        (n for n in wb.sheetnames if n.lower() == "nv cajas (bl)"), None
    )
    if nombre_nv_cajas:
        wb[nombre_nv_cajas].sheet_state = "hidden"

    # ── Hoja Nota de venta ─────────────────────────────────────────────────────
    nombre_hoja = next(
        (n for n in wb.sheetnames if n.lower() == "nota de venta"), None
    )
    if nombre_hoja is None:
        raise ValueError(
            f"No se encontró la hoja 'Nota de venta'.\n"
            f"Hojas disponibles: {', '.join(wb.sheetnames)}"
        )
    ws = wb[nombre_hoja]

    # Datos del cliente
    _fecha_raw = datos_cliente.get("fecha", "")
    try:
        _fecha_dt = datetime.strptime(_fecha_raw, "%d/%m/%Y")
    except Exception:
        _fecha_dt = _fecha_raw
    _set(ws, "C2", _fecha_dt)
    _set(ws, "H1",  datos_cliente.get("cotizacion",   ""))
    _set(ws, "C5",  datos_cliente.get("contacto",     ""))
    _set(ws, "C6",  datos_cliente.get("empresa",      ""))
    _set(ws, "C7",  datos_cliente.get("email",        ""))
    _set(ws, "I5",  datos_cliente.get("razon_social", ""))
    _set(ws, "I6",  datos_cliente.get("rut",          ""))
    _descuento_raw = datos_cliente.get("descuento", "")
    try:
        _descuento_val = float(_descuento_raw) / 100 if _descuento_raw != "" else ""
    except (ValueError, TypeError):
        _descuento_val = _descuento_raw
    _set(ws, "I7",  _descuento_val)
    _set(ws, "C19",  datos_cliente.get("descripcion",   ""))
    _set(ws, "C112", datos_cliente.get("nombre_trabajo",""))
    _set(ws, "C113", datos_cliente.get("descripcion",   ""))

    # Productos (filas 10..17)
    for offset, d in enumerate(datos_productos):
        fila = FILA_INI_PRODUCTOS + offset
        if fila > FILA_FIN_PRODUCTOS:
            break
        _set(ws, f"{COL_PRODUCTO}{fila}", _limpiar_texto(d.get("producto", "")))
        _set(ws, f"{COL_TEXTIL}{fila}",   _limpiar_texto(d.get("textil",   "")))
        _set(ws, f"{COL_ANCHO}{fila}",    d.get("ancho",    ""))
        _set(ws, f"{COL_ALTO}{fila}",     d.get("alto",     ""))
        _set(ws, f"{COL_CANTIDAD}{fila}", d.get("cantidad", ""))
        _set(ws, f"{COL_TEMA}{fila}",     _limpiar_texto(d.get("tema",     "")))

    # Filas sobrantes: limpiar y ocultar
    for fila in range(FILA_INI_PRODUCTOS + len(datos_productos), FILA_FIN_PRODUCTOS + 1):
        for col in [COL_PRODUCTO, COL_TEXTIL, COL_ANCHO, COL_ALTO, COL_CANTIDAD, COL_TEMA]:
            _set(ws, f"{col}{fila}", "")
    _ocultar_sobrantes(ws, FILA_INI_PRODUCTOS, FILA_FIN_PRODUCTOS, len(datos_productos))

    # ── Hoja Cotización ────────────────────────────────────────────────────────
    nombre_cotizacion = next(
        (n for n in wb.sheetnames if n.lower() in ("cotización", "cotizacion")),
        None,
    )
    if nombre_cotizacion:
        ws_cot = wb[nombre_cotizacion]
        _set(ws_cot, "C7", datos_cliente.get("nombre_trabajo", ""))
        _set(ws_cot, "H7", datos_cliente.get("condicion", ""))
        _ocultar_sobrantes(ws_cot, 12, 111, len(datos_productos))

    # ── Hoja OP ───────────────────────────────────────────────────────────────
    nombre_op = next(
        (n for n in wb.sheetnames if n.lower() == "op"),
        None,
    )
    if nombre_op:
        _ocultar_sobrantes(wb[nombre_op], 10, 110, len(datos_productos))

    # ── Abrir en hoja Cotización ───────────────────────────────────────────────
    _activar_hoja_cotizacion(wb)

    wb.save(destino)
    wb.close()
    return str(destino)


def _activar_hoja_cotizacion(wb):
    """Deja la hoja 'Cotización' como hoja activa al abrir el archivo."""
    nombre = next(
        (n for n in wb.sheetnames if n.lower() in ("cotización", "cotizacion")),
        None,
    )
    if nombre:
        wb.active = wb[nombre]


def _ocultar_sobrantes(ws, fila_ini: int, fila_fin: int, n_productos: int):
    """Oculta las filas de producto que no tienen datos."""
    for fila in range(fila_ini + n_productos, fila_fin + 1):
        ws.row_dimensions[fila].hidden = True


def _limpiar_texto(s) -> str:
    """Elimina espacios extremos y capitaliza la primera letra."""
    s = str(s).strip()
    return (s[0].upper() + s[1:]) if s else ""


def _set(ws, celda: str, valor):
    """
    Escribe un valor en una celda. Si la celda pertenece a un rango combinado,
    escribe en la celda superior-izquierda del rango (la única que acepta valores).
    Usa comparación numérica de fila/columna para ser compatible con todas las
    versiones de openpyxl (el operador `in` de MergedCellRange cambió entre versiones).
    """
    from openpyxl.cell import MergedCell
    cell = ws[celda]
    if isinstance(cell, MergedCell):
        row, col = cell.row, cell.column
        for rango in ws.merged_cells.ranges:
            if (rango.min_row <= row <= rango.max_row and
                    rango.min_col <= col <= rango.max_col):
                ws.cell(row=rango.min_row, column=rango.min_col).value = valor
                return
    else:
        cell.value = valor


def _set_number_format(ws, celda: str, fmt: str):
    """
    Aplica un formato de número a una celda, resolviendo celdas combinadas.
    """
    from openpyxl.cell import MergedCell
    cell = ws[celda]
    if isinstance(cell, MergedCell):
        row, col = cell.row, cell.column
        for rango in ws.merged_cells.ranges:
            if (rango.min_row <= row <= rango.max_row and
                    rango.min_col <= col <= rango.max_col):
                ws.cell(row=rango.min_row, column=rango.min_col).number_format = fmt
                return
    else:
        cell.number_format = fmt
